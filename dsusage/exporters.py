"""导出器：Excel (xlsx) / CSV / 官方原始 CSV / meta.json。"""

from __future__ import annotations

import csv
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from .aggregate import ExportTable
from .api import tz_label

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:  # pragma: no cover
    _HAS_OPENPYXL = False

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BAND_FILL = PatternFill("solid", fgColor="EAF1F8")


def ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def make_output_dir(base: Path, start: str, end: str, tz_sec: int) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    name = f"dsu_{start}_{end}_{tz_label(tz_sec).replace(':', '')}_{ts}"
    return ensure_dir(base / name)


def _write_csv(path: Path, table: ExportTable) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=table.columns, extrasaction="ignore")
        writer.writeheader()
        for row in table.rows:
            writer.writerow({c: row.get(c, "") for c in table.columns})


def write_tables_csv(out_dir: Path, tables: List[ExportTable]) -> List[Path]:
    """每个表格写一个 CSV（utf-8-sig，Excel 可直接打开）。"""
    paths: List[Path] = []
    for t in tables:
        p = out_dir / f"{t.name}.csv"
        _write_csv(p, t)
        paths.append(p)
    return paths


def write_xlsx(path: Path, tables: List[ExportTable],
               meta: Optional[Dict[str, Any]] = None) -> None:
    """写多工作表 Excel。"""
    if not _HAS_OPENPYXL:
        raise RuntimeError("未安装 openpyxl，请执行: pip install openpyxl")
    wb = Workbook()
    wb.remove(wb.active)

    if meta:
        ms = wb.create_sheet("导出信息")
        ms.column_dimensions["A"].width = 24
        ms.column_dimensions["B"].width = 60
        for k, v in meta.items():
            ms.append([str(k), str(v)])
        for cell in ms["A"]:
            cell.font = Font(bold=True)

    for t in tables:
        ws = wb.create_sheet(t.title[:31])
        ws.append(t.columns)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for i, row in enumerate(t.rows, start=2):
            ws.append([row.get(c, "") for c in t.columns])
            if i % 2 == 0:
                for cell in ws[i]:
                    cell.fill = BAND_FILL
        for j, col in enumerate(t.columns, start=1):
            letter = get_column_letter(j)
            max_len = len(str(col)) * 2 + 4
            for i in range(2, min(ws.max_row + 1, 200)):
                v = ws.cell(row=i, column=j).value
                if v is not None:
                    max_len = max(max_len, len(str(v)) + 2)
            ws.column_dimensions[letter].width = min(max(max_len, 10), 40)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb.save(path)


def write_raw_csv(out_dir: Path, files: Dict[str, str]) -> List[Path]:
    """把官方导出 zip 内的 CSV 原样保存。"""
    paths: List[Path] = []
    for name, text in files.items():
        safe = Path(name).name
        p = out_dir / ("raw_" + safe)
        p.write_text(text, encoding="utf-8-sig")
        paths.append(p)
    return paths


def write_meta_json(out_dir: Path, meta: Dict[str, Any]) -> Path:
    p = out_dir / "meta.json"
    p.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    return p


def export_all(out_dir: Path, tables: List[ExportTable],
               formats: List[str],
               meta: Optional[Dict[str, Any]] = None,
               raw_csv_files: Optional[Dict[str, str]] = None) -> Dict[str, List[str]]:
    """执行全部导出，返回 {类别: [相对文件名]}。"""
    out_dir = ensure_dir(out_dir)
    result: Dict[str, List[str]] = {"xlsx": [], "csv": [], "raw": [], "meta": []}

    if "xlsx" in formats:
        path = out_dir / "usage.xlsx"
        write_xlsx(path, tables, meta)
        result["xlsx"].append(path.name)
    if "csv" in formats:
        for p in write_tables_csv(out_dir, tables):
            result["csv"].append(p.name)
    if raw_csv_files:
        for p in write_raw_csv(out_dir, raw_csv_files):
            result["raw"].append(p.name)
    if meta:
        result["meta"].append(write_meta_json(out_dir, meta).name)
    return result


def size_human(n: int) -> str:
    for unit in ("B", "KB", "MB", "GB"):
        if n < 1024:
            return f"{n:.1f}{unit}" if unit != "B" else f"{n}B"
        n /= 1024
    return f"{n:.1f}TB"
