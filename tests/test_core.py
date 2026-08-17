"""单元测试：时间工具、API 解析、数据集合并/降级、官方 CSV 解析、聚合、导出。"""

from __future__ import annotations

import json
import os
import sys
import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from dsusage import __version__  # noqa: E402
from dsusage.aggregate import build_tables, compute_totals  # noqa: E402
from dsusage.api import (  # noqa: E402
    DeepSeekPlatformClient, aggregate_to_daily, date_to_end_sec,
    date_to_start_sec, iter_windows, merge_datasets, parse_tz, sec_to_local_dt,
    start_end_sec, tz_label,
)
from dsusage.exporters import export_all, write_xlsx  # noqa: E402
from dsusage.parsing import (  # noqa: E402
    aggregate_official, normalize_utc_date, parse_hour_of_utc_date,
    parse_official_csv,
)

from fixtures import biz_data, make_amount_biz, make_cost_biz  # noqa: E402

TZ8 = 8 * 3600


class TestTimeTools(unittest.TestCase):
    def test_parse_tz(self):
        self.assertEqual(parse_tz("+08:00"), 28800)
        self.assertEqual(parse_tz("-05:30"), -19800)
        self.assertEqual(parse_tz(28800), 28800)
        self.assertEqual(parse_tz(8.0), 28800)
        self.assertEqual(parse_tz("8"), 28800)

    def test_tz_label(self):
        self.assertEqual(tz_label(28800), "UTC+08:00")
        self.assertEqual(tz_label(-19800), "UTC-05:30")
        self.assertEqual(tz_label(0), "UTC+00:00")

    def test_start_end_sec(self):
        s, e = start_end_sec(date(2026, 7, 1), date(2026, 7, 1), TZ8)
        # 北京 2026-07-01 00:00 = UTC 2026-06-30 16:00
        from datetime import timezone
        expect = int(datetime(2026, 6, 30, 16, 0, tzinfo=timezone.utc).timestamp())
        self.assertEqual(s, expect)
        self.assertEqual(e - s, 86400)

    def test_sec_to_local(self):
        s = date_to_start_sec(date(2026, 7, 1), TZ8)
        local = sec_to_local_dt(s, TZ8)
        self.assertEqual(local.strftime("%Y-%m-%d %H:%M"), "2026-07-01 00:00")

    def test_iter_windows(self):
        wins = list(iter_windows(date(2026, 1, 1), date(2026, 3, 1), max_days=30))
        # 60 天 → 2 个 30 天窗口（01-01..01-30 / 01-31..03-01）
        self.assertEqual(len(wins), 2)
        self.assertEqual(wins[0], (date(2026, 1, 1), date(2026, 1, 30)))
        self.assertEqual(wins[1], (date(2026, 1, 31), date(2026, 3, 1)))
        self.assertEqual(wins[-1][1], date(2026, 3, 1))
        # 完整覆盖
        days = set()
        for ws, we in wins:
            d = ws
            while d <= we:
                days.add(d)
                d = d.fromordinal(d.toordinal() + 1)
        total = (date(2026, 3, 1) - date(2026, 1, 1)).days + 1
        self.assertEqual(len(days), total)


class TestParse(unittest.TestCase):
    def _client(self):
        return DeepSeekPlatformClient("test-token")

    def test_parse_amount(self):
        c = self._client()
        s = date_to_start_sec(date(2026, 7, 1), TZ8)
        biz = make_amount_biz(s, s + 3600, 3600, [{
            "api_key": "k1", "model": "deepseek-chat",
            "buckets": [{"time": s, "usage": {"PROMPT_CACHE_HIT_TOKEN": 10,
                                               "PROMPT_CACHE_MISS_TOKEN": 5,
                                               "RESPONSE_TOKEN": 7, "REQUEST": 2}}],
        }])
        ds = c.parse_amount(biz, TZ8, {"k1": "我的Key"})
        self.assertEqual(ds.bucket_sec, 3600)
        self.assertEqual(ds.granularity(), "hourly")
        self.assertEqual(ds.series[0].api_key, "k1")
        self.assertEqual(ds.api_key_names["k1"], "我的Key")
        b = ds.series[0].buckets[0]
        self.assertEqual((b.prompt_cache_hit, b.prompt_cache_miss, b.response, b.request), (10, 5, 7, 2))

    def test_parse_cost_and_merge(self):
        c = self._client()
        s = date_to_start_sec(date(2026, 7, 1), TZ8)
        amt = make_amount_biz(s, s + 3600, 3600, [{
            "api_key": "k1", "model": "deepseek-chat",
            "buckets": [{"time": s, "usage": {"RESPONSE_TOKEN": 100, "REQUEST": 1}}]}])
        cost = make_cost_biz(s, s + 3600, 3600, "CNY", [{
            "api_key": "k1", "model": "deepseek-chat",
            "buckets": [{"time": s, "cost": "0.0123"}]}])
        ds = c.parse_amount(amt, TZ8)
        ds.cost_by_currency = c.parse_cost(cost, TZ8)
        c.merge_cost_into_amount(ds)
        self.assertAlmostEqual(ds.series[0].buckets[0].cost, 0.0123, places=6)
        self.assertIn("CNY", ds.cost_by_currency)

    def test_api_key_as_dict(self):
        """真实平台 cost 响应的 api_key 可能是对象（曾导致 unhashable dict 崩溃）。"""
        c = self._client()
        s = date_to_start_sec(date(2026, 7, 1), TZ8)
        ak_obj = {"tracking_id": "k1", "name": "KeyA", "sensitive_id": "sk-***"}
        amt = make_amount_biz(s, s + 3600, 3600, [{
            "api_key": ak_obj, "model": "deepseek-chat",
            "buckets": [{"time": s, "usage": {"RESPONSE_TOKEN": 100, "REQUEST": 1}}]}])
        cost = make_cost_biz(s, s + 3600, 3600, "CNY", [{
            "api_key": ak_obj, "model": "deepseek-chat",
            "buckets": [{"time": s, "cost": "0.0123"}]}])
        ds = c.parse_amount(amt, TZ8)
        ds.cost_by_currency = c.parse_cost(cost, TZ8)
        c.merge_cost_into_amount(ds)  # 不应再抛 unhashable dict
        self.assertEqual(ds.series[0].api_key, "k1")
        self.assertAlmostEqual(ds.series[0].buckets[0].cost, 0.0123, places=6)
        # 无已知 id 键的 dict 兜底为稳定 JSON
        odd = {"foo": {"bar": 1}}
        self.assertEqual(c._norm_api_key(odd), '{"foo": {"bar": 1}}')

    def test_merge_datasets(self):
        s1 = date_to_start_sec(date(2026, 7, 1), TZ8)
        s2 = date_to_start_sec(date(2026, 7, 2), TZ8)
        c = self._client()
        d1 = c.parse_amount(make_amount_biz(s1, s1 + 86400, 86400, [{
            "api_key": "k1", "model": "m1",
            "buckets": [{"time": s1, "usage": {"RESPONSE_TOKEN": 100, "REQUEST": 1}}]}]), TZ8)
        d2 = c.parse_amount(make_amount_biz(s2, s2 + 86400, 86400, [{
            "api_key": "k1", "model": "m1",
            "buckets": [{"time": s2, "usage": {"RESPONSE_TOKEN": 200, "REQUEST": 2}}]}]), TZ8)
        merged = merge_datasets([d1, d2], TZ8)
        self.assertEqual(len(merged.series), 1)
        self.assertEqual(len(merged.series[0].buckets), 2)
        self.assertEqual(merged.start_sec, s1)
        self.assertEqual(merged.end_sec, s2 + 86400)

    def test_aggregate_to_daily(self):
        c = self._client()
        s = date_to_start_sec(date(2026, 7, 1), TZ8)
        biz = make_amount_biz(s, s + 2 * 3600, 3600, [{
            "api_key": "k1", "model": "m1",
            "buckets": [{"time": s, "usage": {"RESPONSE_TOKEN": 1, "REQUEST": 1}},
                        {"time": s + 3600, "usage": {"RESPONSE_TOKEN": 2, "REQUEST": 1}}]}])
        ds = c.parse_amount(biz, TZ8)
        ds = aggregate_to_daily(ds)
        self.assertEqual(ds.bucket_sec, 86400)
        self.assertEqual(ds.series[0].buckets[0].response, 3)


class TestOfficialCsv(unittest.TestCase):
    CSV = (
        "\ufeffutc_date,model,api_key_name,type,price,amount\n"
        "20260701,deepseek-chat,key-a,input_cache_hit_tokens,0.0000001,1000\n"
        "20260701,deepseek-chat,key-a,input_cache_miss_tokens,0.0000005,500\n"
        "20260701,deepseek-chat,key-a,output_tokens,0.000002,300\n"
        "2026-07-02,deepseek-chat,key-a,request_count,0.0,10\n"
    )

    def test_normalize(self):
        self.assertEqual(normalize_utc_date("20260701"), "2026-07-01")
        self.assertEqual(normalize_utc_date("2026-07-01"), "2026-07-01")
        self.assertEqual(parse_hour_of_utc_date("2026-07-01 03:00:00"), 3)
        self.assertIsNone(parse_hour_of_utc_date("2026-07-01"))

    def test_parse_and_aggregate(self):
        rows = parse_official_csv(self.CSV)
        self.assertEqual(len(rows), 4)
        agg = aggregate_official(rows)
        self.assertEqual(agg.totals["cache_hit"], 1000)
        self.assertEqual(agg.totals["cache_miss"], 500)
        self.assertEqual(agg.totals["response"], 300)
        self.assertEqual(agg.totals["requests"], 10)
        self.assertFalse(agg.has_hourly())
        self.assertEqual(len(agg.daily_summary), 2)

    def test_hourly_csv(self):
        csv = (
            "utc_date,model,api_key_name,type,price,amount\n"
            "2026-07-01 00:00:00,deepseek-chat,key-a,output_tokens,0.000002,100\n"
            "2026-07-01 01:00:00,deepseek-chat,key-a,output_tokens,0.000002,200\n"
        )
        rows = parse_official_csv(csv)
        agg = aggregate_official(rows)
        self.assertTrue(agg.has_hourly())
        self.assertEqual(len(agg.hourly), 2)
        self.assertEqual(agg.hourly[1]["hour"], 1)


class TestAggregate(unittest.TestCase):
    def _ds(self):
        c = DeepSeekPlatformClient("t")
        s1 = date_to_start_sec(date(2026, 7, 1), TZ8)
        s2 = date_to_start_sec(date(2026, 7, 2), TZ8)
        amt = make_amount_biz(s1, s2 + 86400, 86400, [
            {"api_key": "k1", "model": "deepseek-chat",
             "buckets": [{"time": s1, "usage": {"PROMPT_CACHE_HIT_TOKEN": 10,
                                                 "RESPONSE_TOKEN": 5, "REQUEST": 1}}]},
            {"api_key": "k2", "model": "deepseek-reasoner",
             "buckets": [{"time": s2, "usage": {"PROMPT_CACHE_MISS_TOKEN": 20,
                                                 "RESPONSE_TOKEN": 8, "REQUEST": 2}}]},
        ])
        cost = make_cost_biz(s1, s2 + 86400, 86400, "CNY", [
            {"api_key": "k1", "model": "deepseek-chat",
             "buckets": [{"time": s1, "cost": "0.01"}]},
            {"api_key": "k2", "model": "deepseek-reasoner",
             "buckets": [{"time": s2, "cost": "0.02"}]},
        ])
        ds = c.parse_amount(amt, TZ8, {"k1": "KeyA", "k2": "KeyB"})
        ds.cost_by_currency = c.parse_cost(cost, TZ8)
        c.merge_cost_into_amount(ds)
        return ds

    def test_build_tables(self):
        ds = self._ds()
        tables = build_tables(ds)
        names = [t.name for t in tables]
        self.assertIn("daily_detail", names)
        self.assertIn("daily_summary", names)
        self.assertIn("model_summary", names)
        self.assertIn("api_key_summary", names)
        self.assertIn("cost_detail", names)
        # 模型汇总两行
        model_t = next(t for t in tables if t.name == "model_summary")
        self.assertEqual(len(model_t.rows), 2)

    def test_compute_totals(self):
        t = compute_totals(self._ds())
        self.assertEqual(t["requests"], 3)
        self.assertEqual(t["cache_hit"], 10)
        self.assertEqual(t["cache_miss"], 20)
        self.assertEqual(t["response"], 13)
        self.assertEqual(t["total_tokens"], 43)
        self.assertAlmostEqual(t["cost"], 0.03, places=6)


class TestExporters(unittest.TestCase):
    def test_export_all(self):
        from dsusage.aggregate import ExportTable
        t = ExportTable(name="daily_summary", title="每日汇总",
                        columns=["日期", "请求数"],
                        rows=[{"日期": "2026-07-01", "请求数": 3}])
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "out"
            result = export_all(out, [t], ["xlsx", "csv"],
                                meta={"k": "v"},
                                raw_csv_files={"amount-1.csv": "a,b\n1,2\n"})
            self.assertIn("usage.xlsx", result["xlsx"])
            self.assertIn("daily_summary.csv", result["csv"])
            self.assertIn("raw_amount-1.csv", result["raw"])
            self.assertIn("meta.json", result["meta"])
            self.assertTrue((out / "usage.xlsx").exists())
            # 校验 xlsx 可读
            from openpyxl import load_workbook
            wb = load_workbook(out / "usage.xlsx")
            self.assertIn("每日汇总", wb.sheetnames)


class TestCli(unittest.TestCase):
    def test_version(self):
        self.assertEqual(__version__, "1.0.3")

    def test_parser_builds(self):
        from dsusage.cli import build_parser
        p = build_parser()
        args = p.parse_args(["range", "--start", "2026-07-01", "--end", "2026-07-05"])
        self.assertEqual(args.cmd, "range")
        self.assertEqual(args.granularity, "auto")


if __name__ == "__main__":
    unittest.main()
